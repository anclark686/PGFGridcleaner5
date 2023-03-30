import os
from datetime import datetime, date, time, timedelta
import json
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border, colors
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from operator import itemgetter
import re
from components.helpers import is_dst, make_pretty_notes


desktop = os.path.join(os.path.join(os.path.expanduser("~")), "Desktop")
gm_path = f"{desktop}/GridMaster"

with open("./json_templates/grid_templates.json") as grid_templates_file:
    grid_templates = json.load(grid_templates_file)

# this is a template file, not project specific
with open("../frontend/src/json_templates/projectData.json") as project_data:
    project_templates = json.load(project_data)


def check_master(ws, master):
    m_spot = ws["c4"].value
    if type(m_spot) != int:
        m_spot = ws["c4"].value.split(" ")[0].lower()
        if m_spot[0] != "m":
            m_spot = "m" + m_spot
    else:
        m_spot = "m" + str(m_spot)

    print(master)
    print(m_spot)
    if master != m_spot:
        return False
    else:
        return True


def get_raw_data(ws):
    pop_rows = []

    for row in ws.iter_rows(8):
        row_vals = []
        none_count = 0
        found = False
        color = ""
        for cell in row[1:11]:
            row_vals.append(cell.value)
            cell_color = cell.fill.start_color.index
            if cell_color != 00000000 or cell_color != 0:
                color = cell_color
            if cell.value == None:
                none_count += 1
            if cell.value == "* required information":
                found = True
                break
        if none_count <= 6 and len(row_vals) > 1:
            row_vals.append(color)
            pop_rows.append(row_vals)
        if found:
            break
    return pop_rows


def clean_date(messy_date, days, months, suffixes):
    today = date.today()
    current_year = str(today.year)
    new_date = ""
    date_data1 = messy_date.replace(
        ",", "").replace("-", " ").replace("/", " ").strip().split(" ")
    no_days = [x for x in date_data1 if x.lower() not in days]
    no_months = [grid_templates["months"][x.lower()] if x.lower()
                 in months else x for x in no_days]
    no_suffixes = [x[:-2] if x[-2:] in suffixes else x for x in no_months]

    if len(no_suffixes) < 3:

        if today.month < 11:
            no_suffixes.append(current_year)
        else:
            next_year = str(int(current_year) + 1)
            if int(no_suffixes[0]) < 3:
                no_suffixes.append(next_year)
            else:
                no_suffixes.append(current_year)

    try:
        new_date = datetime.strptime(("/").join(no_suffixes), '%m/%d/%Y')
    except ValueError as err:
        try:
            new_date = datetime.strptime(("/").join(no_suffixes), '%m/%d/%y')
        except ValueError as err:
            new_date - "UNABLE TO DETERMINE DATE"
    return new_date


def clean_time(messy_time, rdate, iv_or_rsp, country, tzones=None):
    time_list = ["", ""]
    messy_time = messy_time.replace(";", ":")
    abc_time = []
    am_pm = ""
    military = False
    potential_tz = ""
    new_time = None
    issue = ""

    if messy_time == "None":
        issue = "CHECK TIME W/ RCR"
        messy_time = "1200"

    if messy_time == "Noon":
        messy_time = "1200"
        am_pm = "pm"

    if "-" in messy_time or "to" in messy_time:
        if "-" in messy_time:
            times = messy_time.split("-")
        else:
            times = messy_time.split("to")
        times = [x.strip() for x in times]
        if "am" in times[1] and "am" not in times[0]:
            times[0] += "am"
            messy_time = times[0]
        elif "pm" in times[1] and "pm" not in times[0]:
            times[0] += "pm"
            messy_time = times[0]
        else:
            messy_time = times[0]

    for i in messy_time:
        if i.isnumeric() or i == ":":
            time_list[0] += i
        else:
            time_list[1] += i.lower()

    time_list[0] = time_list[0].replace(":", "")
    if len(time_list[0]) <= 2:
        time_list[0] += "00"
    if len(time_list[0]) == 3:
        time_list[0] = "0" + time_list[0]
    if int(time_list[0][0:2]) > 12:
        military = True

    time_list[1] = time_list[1].replace(".", "").replace(",", "").strip()
    abc_time = time_list[1].split()
    if len(abc_time) >= 1:
        if "am" in abc_time[0] or "a" in abc_time[0]:
            am_pm = "am"
        elif "pm" in abc_time[0] or "p" in abc_time[0]:
            am_pm = "pm"

    if len(abc_time) > 1:
        if am_pm != "":
            potential_tz = " ".join(abc_time[1:])
        else:
            potential_tz = " ".join(abc_time)

    if potential_tz != "" and iv_or_rsp == "iv":
        if potential_tz in tzones.keys():
            col_tz = tzones[potential_tz]
            if col_tz != "US/Eastern":
                # do something to change to eastern time
                pass
        potential_tz = ""

    if am_pm == "" and military == False:
        if (country == "United States" or country == "Canada"):
            if int(time_list[0][0:2]) == 12:
                am_pm = "pm"
            elif int(time_list[0][0:2]) < 6:
                am_pm = "pm"
            elif int(time_list[0][0:2]) > 9 and int(time_list[0][0:2]) < 12:
                am_pm = "am"
            else:
                issue = f"CHECK AM OR PM {iv_or_rsp}"
                am_pm = "am"
        else:
            issue = f"CHECK AM OR PM {iv_or_rsp}"
            am_pm = "am"
    # print(issue)
    if military:
        new_time = datetime.strptime(time_list[0], "%H%M").time()
        new_time = datetime.combine(rdate, new_time)
    else:
        new_time = datetime.strptime(time_list[0] + am_pm, "%I%M%p").time()
        new_time = datetime.combine(rdate, new_time)
    return [new_time, potential_tz, issue]


def find_timezone(time1, time2, tz_country, potential_tz, tzones, offsets, country):
    time_diff = None

    # print(f"time1: {time1}, time2: {time2}, tz_col: {tz_country}, potential: {potential_tz}")
    # print(tzones)
    if ((type(time1) == str or type(time2) == str) and
            tz_country == "None" and potential_tz == ""):
        return "UNABLE TO DETERMINE"

    if potential_tz.lower() in tzones.keys():
        return tzones[potential_tz.lower()]
    elif tz_country.lower() in tzones.keys():
        return tzones[tz_country.lower()]
    else:
        if country != "United States" or country != "Canada":
            if country.lower() in tzones.keys():
                return tzones[country.lower()]

    if (type(time1) != str and type(time2) != str):
        if time1 == time2:
            time_diff = 0.0
        elif time1 > time2:
            time_diff = (time1 - time2).total_seconds() / (60*60)
            time_diff = -time_diff
        else:
            time_diff = (time2 - time1).total_seconds() / (60*60)
        if str(time_diff) in offsets.keys():
            return offsets[str(time_diff)]
    return "UNABLE TO DETERMINE"


# def final_time_comparison(time1, time2, timezone):
#     # print("we in here")
#     # chec to see if all the iv times are wrong, may need to adjust tz
#     return [time1, time2]


def format_names(rsp_name, invite_name, display_name, prefixes):
    if invite_name == "None":
        invite_name = rsp_name

    if display_name == "None":
        if "," in rsp_name:
            full_name = rsp_name.split(",")
            last = full_name[0]
            full_name.append(last)
            full_name.remove(last)
        else:
            full_name = rsp_name.split(" ")
        if len(full_name) >= 2 and full_name[0].lower() not in prefixes:
            display_name = full_name[0]
        elif len(full_name) >= 2 and full_name[0].lower() in prefixes:
            if len(full_name[1]) > 2 or full_name[1][-1] != ".":
                display_name = f"{full_name[0]} {full_name[1][0]}."
            elif len(full_name[1]) == 1:
                display_name = f"{full_name[0]} {full_name[1]}"
        else:
            display_name = full_name

    return [rsp_name, invite_name, display_name]


def check_phones(phone1, phone2, t_zone, abbrev, pretty_tzs, sess_type, country):
    def check_if_invalid(phone):
        invalid_terms = ["n/a", "n.a", "n.a.", "none", "-", "--"]
        for i in invalid_terms:
            if phone.lower() == i:
                return ""
        return phone

    phone1 = check_if_invalid(phone1)
    phone2 = check_if_invalid(phone2)

    if phone1 == "" and phone2 == "":
        return [phone1, phone2]
    elif phone1 == "" and phone2 != "":
        phone1, phone2 = [phone2, phone1]

    if sess_type != "IDI":
        pretty_tz = ""
        if country != "United States" or country != "Canada":
            if country in abbrev.keys():
                pretty_tz = abbrev[country]
        elif t_zone != "UNABLE TO DETERMINE":
            if t_zone in pretty_tzs.keys():
                pretty_tz = pretty_tzs[t_zone]
        phone1 = f"{pretty_tz} {phone1}" if (
            pretty_tz != "" and pretty_tz not in phone1) else phone1
    return [phone1, phone2]


def check_email(email):
    regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'

    if (re.fullmatch(regex, email)):
        return True
    else:
        return False


def check_dup_emails(email_list):
    print(email_list)
    email_set = set(email_list)
    if len(email_set) == 1:
        return email_list[0]
    elif len(email_set) != len(email_list):
        return "Duplicate emails found."
    return ""


def find_duplicates(rsp_list):
    for i in range(1,len(rsp_list)):
        prev_date, prev_time = rsp_list[i-1]["date"], rsp_list[i-1]["iv_time"]
        prev_date = datetime.combine(prev_date, prev_time)
        curr_date, curr_time = rsp_list[i]["date"], rsp_list[i]["iv_time"]
        current_date = datetime.combine(curr_date, curr_time)
        if current_date == prev_date:
            prev_date_bts = prev_date.strftime("%#m/%#d/%#y %#I:%M %p")
            prev_rsp = rsp_list[i-1]["rsp_name"]
            rsp_list[i]["issues"].append(f"Potential Overlap with: {prev_date_bts} {prev_rsp}")
    return rsp_list


def gathernotes(row_dict):
    pass


def find_start_end(row_list):
    given_date = date(2000, 1, 1)
    start_time = time(0, 0, 0)
    end_time = time(0, 0, 0)
    times = []
    for i in row_list:
        row_date = i["date"]
        row_time = i["iv_time"]

        if row_date != given_date:
            if start_time != time(0, 0, 0):
                pretty_start = start_time.strftime("%#I:%M%p")
                ugly_end = datetime.combine(datetime.today(), end_time)
                ugly_end = (ugly_end + timedelta(hours=1)).time()
                pretty_end = (ugly_end).strftime("%#I:%M%p")
                times.append(f"{pretty_start}-{pretty_end}")
            given_date = row_date
            start_time = row_time
            end_time = row_time
        else:
            if row_time < start_time:
                print("this shouldn't happen")
            if row_time > end_time:
                end_time = row_time
        if i == row_list[-1]:
            pretty_start = start_time.strftime("%#I:%M%p")
            ugly_end = datetime.combine(datetime.today(), end_time)
            ugly_end = (ugly_end + timedelta(hours=1)).time()
            pretty_end = (ugly_end).strftime("%#I:%M%p")
            times.append(f"{pretty_start}-{pretty_end}")
    return times


def save_rsp_grid_details(row_list, dates, times, spoc):
    json_data = {
        "rsp_data": row_list,
        "dates": dates,
        "times": times,
        "spoc": spoc
    }
    with open("./json_templates/rsp_grid_details.json", "w") as outfile:
        json.dump(json_data, outfile, default=str)


def populate_cleaned_data(ws, project_details):
    days = grid_templates["days"].keys()
    months = grid_templates["months"].keys()
    suffixes = grid_templates["suffixes"]
    abbrev = project_templates["Abbreviations"]
    pretty_tzs = grid_templates["pretty-tz"]
    prefixes = grid_templates["prefixes"]
    colors = grid_templates["grid_colors"]

    sess_type = project_details["session-type"]
    country = project_details["country"]

    rows = get_raw_data(ws)
    row_list = []
    dates = []
    times = []
    notes = []
    emails = []

    for i in rows:
        row_dict = {}
        rsp_date = i[0]
        iv_time = i[1]
        rsp_time = i[2]
        tz_country = str(i[3]).strip()
        rsp_name = str(i[4]).strip()
        invite_name = str(i[5]).strip()
        display_name = str(i[6]).strip()
        phone1 = str(i[7]) if str(i[9]) != "None" else ""
        phone2 = str(i[8]) if str(i[8]) != "None" else ""
        email = str(i[9]).strip()
        row_color_hex = str(i[10])
        time_issue = ""

        badness = []
        potential_tz = ""

        if not isinstance(rsp_date, datetime):
            rsp_date = clean_date(rsp_date, days, months, suffixes)
        if rsp_date not in dates:
            dates.append(rsp_date)

        dst = is_dst(rsp_date, "US/Eastern")
        if dst:
            offsets = grid_templates["dt-time-differences"]
            tzones = grid_templates["DT_TZ_SHORT"]
        else:
            offsets = grid_templates["st-time-differences"]
            tzones = grid_templates["ST_TZ_SHORT"]

        if not isinstance(iv_time, time):
            iv_time, null, time_issue = clean_time(
                str(iv_time), rsp_date, "iv", country, tzones)
        else:
            iv_time = datetime.combine(rsp_date, iv_time)
        if time_issue:
            badness.append(time_issue)

        if not isinstance(rsp_time, time):
            if rsp_time == None and not time_issue:
                rsp_time = iv_time
            else:
                rsp_time, potential_tz, time_issue = clean_time(
                    str(rsp_time), rsp_date, "rsp", country)
        else:
            rsp_time = datetime.combine(rsp_date, rsp_time)
        if time_issue:
            badness.append(time_issue)

        tz_country = find_timezone(
            iv_time, rsp_time, tz_country, potential_tz, tzones, offsets, country)
        tz_time_diff = find_timezone(
            iv_time, rsp_time, "", "", tzones, offsets, country)

        if tz_country != tz_time_diff:
            tz_country = "TIMEZONE MISMATCH"
            badness.append("TIMEZONE MISMATCH")

        # if tz_country != "UNABLE TO DETERMINE":
        #     if isinstance(iv_time, datetime) and isinstance(rsp_time, datetime):
        #         iv_time, rsp_time = final_time_comparison(iv_time, rsp_time, tz_country)

        rsp_date = rsp_date.date()
        iv_time = iv_time.time() if isinstance(iv_time, datetime) else iv_time
        rsp_time = rsp_time.time() if isinstance(rsp_time, datetime) else rsp_time

        rsp_name, invite_name, display_name = format_names(
            rsp_name, invite_name, display_name, prefixes)

        phone1, phone2 = check_phones(
            phone1, phone2, tz_country, abbrev, pretty_tzs, sess_type, country)

        valid_email = check_email(email)
        if not valid_email:
            badness.append("BAD EMAIL")
        else:
            emails.append(email)

        if row_color_hex == "0" or row_color_hex == "00000000":
            row_color_hex = "FFFFFFFF"
        elif row_color_hex == "5":
            row_color_hex = "FFFF0000"
        elif row_color_hex == "6":
            row_color_hex = "FF00FF00"

        if row_color_hex not in colors.keys():
            badness.append("UNABLE TO DETERMINE COLOR")
            row_color = "strange"
        else:
            row_color = colors[row_color_hex]


        row_dict = {
            "date": rsp_date,
            "iv_time": iv_time,
            "rsp_time": rsp_time,
            "time_zone": tz_country,
            "rsp_name": rsp_name,
            "invite_name": invite_name,
            "display_name": display_name,
            "phone1": phone1,
            "phone2": phone2,
            "email": email,
            "row_color_hex": row_color_hex,
            "row_color": row_color,
            "issues": badness,
        }
        row_list.append(row_dict)

    sorted_rows = sorted(row_list, key=itemgetter(
        'date', 'iv_time', 'display_name'))
    if sess_type == "IDI":
        sorted_rows = find_duplicates(sorted_rows)


    for i in sorted_rows:
        if i["issues"]:
            pretty_notes = make_pretty_notes(i["issues"])
            idate, itime = i["date"], i["iv_time"]
            dt_combo = datetime.combine(idate, itime)
            str_datetime = datetime.strftime(dt_combo, "%#m/%#d/%#y %#I:%M %p")
            name = i["display_name"]
            note_str = f"{str_datetime} - {name}: {pretty_notes}"
            print(note_str)
            notes.append(note_str)

    dates = sorted(dates)
    dates = [d.strftime("%#m/%#d/%Y") for d in dates]
    times = find_start_end(sorted_rows)
    spoc = check_dup_emails(emails)
    if spoc == "Duplicate emails found.":
        notes.append(spoc)

    save_rsp_grid_details(sorted_rows, dates, times, spoc)
    
    return [sorted_rows, dates, times, notes]


def make_new_excel(row_list):
    new_wb = openpyxl.Workbook()
    sheet = new_wb.active
    dim_holder = DimensionHolder(worksheet=sheet)

    for i in range(len(row_list)):
        value = row_list[i]
        i += 1
        row_color_hex = PatternFill(start_color=value["row_color_hex"],
                                    end_color=value["row_color_hex"],
                                    fill_type='solid')

        something_wrong = PatternFill(start_color="1CC9FFFF",
                                      end_color="1CC9FFFF",
                                      fill_type='solid')

        found_cell = sheet.cell(row=i, column=1)
        if "not found" not in value["project_nums"] and value["project_nums"]:
            found_note = ""
            for x in value["project_nums"]:
                if "Scheduled" in list(x.values())[0]:
                    found_note += f"Found on {list(x.keys())[0]}. "
                elif "Deleted" in list(x.values())[0]:
                    found_note += f"Prev. deleted from {list(x.keys())[0]}. "
                elif "Rescheduled" in list(x.values())[0]:
                    found_note += f"Prev. scheduled on {list(x.keys())[0]}. "
            found_cell.value = found_note
                
        else:
            found_cell.value = "New"                          

        date_cell = sheet.cell(row=i, column=2)
        date_cell.value = value["date"]
        date_cell.fill = row_color_hex

        iv_time_cell = sheet.cell(row=i, column=3)
        iv = value["iv_time"]
        iv_time_cell.value = iv if type(
            iv) == str else iv.strftime("%#I:%M %p")
        iv_time_cell.fill = row_color_hex if "CHECK AM OR PM iv" not in value[
            "issues"] else something_wrong

        rsp_time_cell = sheet.cell(row=i, column=4)
        rs = value["rsp_time"]
        rsp_time_cell.value = rs if type(
            rs) == str else rs.strftime("%#I:%M %p")
        rsp_time_cell.fill = row_color_hex if "CHECK AM OR PM rsp" not in value[
            "issues"] else something_wrong

        tz_cell = sheet.cell(row=i, column=5)
        tz_cell.value = value["time_zone"]
        tz_cell.fill = row_color_hex if "UNABLE" not in value[
            "time_zone"] or "MISMATCH" not in value["time_zone"] else something_wrong

        r_name_cell = sheet.cell(row=i, column=6)
        r_name_cell.value = value["rsp_name"]
        r_name_cell.fill = row_color_hex

        i_name_cell = sheet.cell(row=i, column=7)
        i_name_cell.value = value["invite_name"]
        i_name_cell.fill = row_color_hex

        d_name_cell = sheet.cell(row=i, column=8)
        d_name_cell.value = value["display_name"]
        d_name_cell.fill = row_color_hex

        phone1_cell = sheet.cell(row=i, column=9)
        phone1_cell.value = value["phone1"]
        phone1_cell.fill = row_color_hex

        phone2_cell = sheet.cell(row=i, column=10)
        phone2_cell.value = value["phone2"]
        phone2_cell.fill = row_color_hex

        email_cell = sheet.cell(row=i, column=11)
        email_cell.value = value["email"]
        email_cell.fill = row_color_hex if "BAD EMAIL" not in value["issues"] else something_wrong

        if len(value["issues"]) >= 1 or "stale" in value["row_color"]:
            pretty_notes = make_pretty_notes(value["issues"])

            if "stale" in value["row_color"]:
                pretty_notes += "Stale yellow, already on PC. "

            notes_cell = sheet.cell(row=i, column=12)
            notes_cell.value = pretty_notes
            notes_cell.fill = something_wrong

        for col in range(sheet.min_column, sheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(
                sheet, min=col, max=col, width=20)

        sheet.column_dimensions = dim_holder

        new_wb.save("newFile.xlsx")


def clean_and_check_grid(file_path, master_num, project_details):
    wb = load_workbook(file_path)
    ws = wb.active

    notes = []

    good_master = check_master(ws, master_num)
    if not good_master:
        notes.append("Master mismatch")

    row_list, grid_dates, grid_times, grid_notes = populate_cleaned_data(
        ws, project_details)
    notes.extend(grid_notes)
    print(notes)
    # make_new_excel(row_list)

    return grid_dates, grid_times, notes

# populate_cleaned_data()


info_dict = {
    "master": "m146337",
    "fac": "InterVu (Europe)",
    "country": "Germany",
    "language": "German",
    "two-channel": "220",
    "session-type": "IDI",
    "webcam": "WebCam",
    "project-notes": "test project\nRef: 12345 | DE Webcam IDIs | NO TS | 2CA German",
    "project-numbers": ["600612", "600352", "600353"],
    "dates": ["1/1/2023", "12/16/2023", "12/17/2023"],
    "times": ["9:00AM-6:00PM", "7:00AM-4:00PM", "6:00AM-7:00PM"],
    "statuses": ["Active", "Active", "Active"],
    "grid-dates": [
        "2/15/2021",
        "2/16/2021",
        "2/17/2021",
        "2/18/2021",
        "2/23/2021",
        "2/24/2021",
        "2/25/2021",
        "3/1/2021",
        "8/13/2022",
        "8/14/2022",
        "8/15/2022"
    ],
    "grid-times": [
        "2:15PM-5:45PM",
        "6:00AM-5:45PM",
        "6:00AM-5:00PM",
        "10:00PM-11:00PM",
        "8:45AM-9:45AM",
        "8:45AM-4:30PM",
        "8:15AM-4:00PM",
        "3:00PM-4:00PM",
        "1:30PM-6:15PM",
        "8:00AM-6:00PM",
        "2:00PM-4:45PM"
    ]
}
# clean_and_check_grid(os.path.join(gm_path, "m146337.xlsx"), "m146337", info_dict)
